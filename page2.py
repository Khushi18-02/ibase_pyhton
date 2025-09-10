import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os
import subprocess
import platform
from datetime import datetime

FILE_NAME = "med2.xlsx"
MAX_GENERICS = 5

COLORS = {
    'primary': '#2E86AB',
    'secondary': '#A23B72',
    'accent': '#F18F01',
    'success': '#28A745',
    'background': '#F5F7FA',
    'card': '#FFFFFF',
    'text_dark': '#2C3E50',
    'text_light': '#7F8C8D'
}

class ModernButton:
    def __init__(self, parent, text, command, bg_color, hover_color=None, width=20):
        self.bg_color = bg_color
        self.hover_color = hover_color or self._lighten_color(bg_color)
        
        self.button = tk.Button(
            parent, 
            text=text, 
            command=command,
            bg=bg_color,
            fg='white',
            font=('Segoe UI', 11, 'bold'),
            relief='flat',
            cursor='hand2',
            padx=20,
            pady=8,
            width=width,
            borderwidth=0
        )
        
        self.button.bind('<Enter>', self._on_enter)
        self.button.bind('<Leave>', self._on_leave)
    
    def _lighten_color(self, color):
        colors = {
            '#2E86AB': '#3A9BC1',
            '#A23B72': '#B8457E', 
            '#F18F01': '#FFA01E',
            '#28A745': '#4CD964'
        }
        return colors.get(color, color)
    
    def _on_enter(self, event):
        self.button.config(bg=self.hover_color)
    
    def _on_leave(self, event):
        self.button.config(bg=self.bg_color)
    
    def pack(self, **kwargs):
        self.button.pack(**kwargs)
    
    def grid(self, **kwargs):
        self.button.grid(**kwargs)

def initialize_excel():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Medicine Name", "Composition", "Date Added",
            "Generic 1 Name", "Generic 1 Composition", "Generic 1 Price", "Generic 1 Side Effects",
            "Generic 2 Name", "Generic 2 Composition", "Generic 2 Price", "Generic 2 Side Effects",
            "Generic 3 Name", "Generic 3 Composition", "Generic 3 Price", "Generic 3 Side Effects",
            "Generic 4 Name", "Generic 4 Composition", "Generic 4 Price", "Generic 4 Side Effects",
            "Generic 5 Name", "Generic 5 Composition", "Generic 5 Price", "Generic 5 Side Effects"
        ])
        wb.save(FILE_NAME)

def view_database():
    try:
        if platform.system() == "Windows":
            os.startfile(FILE_NAME)
        elif platform.system() == "Darwin":
            subprocess.call(["open", FILE_NAME])
        else:
            subprocess.call(["xdg-open", FILE_NAME])
    except Exception as e:
        messagebox.showerror("Error", f"Error opening file: {str(e)}")

class MedicineApp:
    def __init__(self):
        self.root = tk.Tk()
        self.setup_main_window()
        self.create_main_interface()
        
    def setup_main_window(self):
        self.root.title("💊 Medicine Management System")
        self.root.geometry("900x700")
        self.root.configure(bg=COLORS['background'])
        self.root.resizable(True, True)
        self.style = ttk.Style()
        self.style.theme_use('clam')
        
    def create_main_interface(self):
        header_frame = tk.Frame(self.root, bg=COLORS['primary'], height=80)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)
        
        tk.Label(
            header_frame,
            text="💊 Medicine Management System",
            font=('Segoe UI', 24, 'bold'),
            fg='white',
            bg=COLORS['primary']
        ).pack(expand=True)
        
        main_frame = tk.Frame(self.root, bg=COLORS['background'])
        main_frame.pack(fill='both', expand=True, padx=30, pady=30)
        
        search_card = tk.Frame(main_frame, bg=COLORS['card'], relief='raised', bd=2)
        search_card.pack(fill='x', pady=(0, 20))
        
        tk.Label(
            search_card,
            text="🔍 Search Medicine",
            font=('Segoe UI', 18, 'bold'),
            fg=COLORS['text_dark'],
            bg=COLORS['card']
        ).pack(pady=(15, 10))
        
        search_frame = tk.Frame(search_card, bg=COLORS['card'])
        search_frame.pack(pady=(0, 15))
        
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(
            search_frame,
            textvariable=self.search_var,
            font=('Segoe UI', 14),
            width=30,
            relief='flat',
            bd=5
        )
        self.search_entry.pack(side='left', padx=(0, 10))
        self.search_entry.bind('<Return>', lambda e: self.search_medicine())
        
        ModernButton(search_frame, "🔍 Search", self.search_medicine, COLORS['primary'], width=12).pack(side='left')
        
        self.search_status = tk.Label(search_card, text="", font=('Segoe UI', 10), bg=COLORS['card'])
        self.search_status.pack(pady=(0, 15))
        
        action_frame = tk.Frame(main_frame, bg=COLORS['background'])
        action_frame.pack(fill='x', pady=10)
        
        ModernButton(action_frame, "➕ Add New Medicine", self.open_medicine_ui, COLORS['success'], width=18).pack(side='left', padx=(0, 10))
        ModernButton(action_frame, "📊 View Database", view_database, COLORS['secondary'], width=18).pack(side='left', padx=(0, 10))
        ModernButton(action_frame, "📈 Statistics", self.show_statistics, COLORS['accent'], width=18).pack(side='left')
        
        self.create_recent_section(main_frame)
        
    def create_recent_section(self, parent):
        recent_card = tk.Frame(parent, bg=COLORS['card'], relief='raised', bd=2)
        recent_card.pack(fill='both', expand=True, pady=(20, 0))
        
        tk.Label(
            recent_card,
            text="📋 Recent Medicines",
            font=('Segoe UI', 16, 'bold'),
            fg=COLORS['text_dark'],
            bg=COLORS['card']
        ).pack(pady=(15, 10))
        
        canvas = tk.Canvas(recent_card, bg=COLORS['card'], height=200)
        scrollbar = ttk.Scrollbar(recent_card, orient='vertical', command=canvas.yview)
        self.recent_frame = tk.Frame(canvas, bg=COLORS['card'])
        self.recent_frame.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        
        canvas.create_window((0, 0), window=self.recent_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side='left', fill='both', expand=True, padx=15, pady=(0, 15))
        scrollbar.pack(side='right', fill='y', pady=(0, 15))
        
        self.load_recent_medicines()
        
    def load_recent_medicines(self):
        try:
            wb = load_workbook(FILE_NAME)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            
            if not rows:
                tk.Label(self.recent_frame, text="No medicines added yet.", font=('Segoe UI', 11), fg=COLORS['text_light'], bg=COLORS['card']).pack(pady=20)
                wb.close()
                return
            
            for row in reversed(rows[-5:]):
                frame = tk.Frame(self.recent_frame, bg='#F8F9FA', relief='raised', bd=1)
                frame.pack(fill='x', pady=2, padx=5)
                
                tk.Label(frame, text=f"📝 {row[0]}", font=('Segoe UI', 12, 'bold'), fg=COLORS['text_dark'], bg='#F8F9FA').pack(anchor='w', padx=10, pady=(5,0))
                tk.Label(frame, text=f"Composition: {row[1][:50]}..." if len(str(row[1])) > 50 else f"Composition: {row[1]}", font=('Segoe UI', 10), fg=COLORS['text_light'], bg='#F8F9FA').pack(anchor='w', padx=10, pady=(0,5))
            
            wb.close()
        except Exception as e:
            tk.Label(self.recent_frame, text="Error loading recent medicines", font=('Segoe UI', 11), fg='red', bg=COLORS['card']).pack(pady=20)
    
    def search_medicine(self):
        search_name = self.search_var.get().strip()
        if not search_name:
            self.search_status.config(text="⚠️ Please enter a medicine name!", fg='red')
            return
        
        try:
            wb = load_workbook(FILE_NAME)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[0].lower() == search_name.lower():
                    wb.close()
                    self.display_medicine(row)
                    self.search_status.config(text="✅ Medicine found!", fg='green')
                    return
            wb.close()
            self.search_status.config(text="❌ Medicine not found. Opening add form...", fg='orange')
            self.root.after(1000, self.open_medicine_ui)
        except Exception as e:
            self.search_status.config(text=f"❌ Error: {str(e)}", fg='red')
    
    def display_medicine(self, data):
        window = tk.Toplevel(self.root)
        window.title(f"💊 {data[0]} - Details")
        window.geometry("900x700")
        window.configure(bg=COLORS['background'])
        window.transient(self.root)
        window.grab_set()
        
        header_frame = tk.Frame(window, bg=COLORS['primary'], height=60)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)
        tk.Label(header_frame, text=f"💊 {data[0]}", font=('Segoe UI', 20, 'bold'), fg='white', bg=COLORS['primary']).pack(expand=True)
        
        content_frame = tk.Frame(window, bg=COLORS['background'])
        content_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        info_card = tk.Frame(content_frame, bg=COLORS['card'], relief='raised', bd=2)
        info_card.pack(fill='x', pady=(0,15))
        tk.Label(info_card, text="Medicine Information", font=('Segoe UI', 16, 'bold'), fg=COLORS['text_dark'], bg=COLORS['card']).pack(pady=(15,10))
        tk.Label(info_card, text=f"Composition: {data[1]}", font=('Segoe UI', 12), fg=COLORS['text_dark'], bg=COLORS['card'], wraplength=800).pack(pady=(0,15), padx=20)
        
        generics_card = tk.Frame(content_frame, bg=COLORS['card'], relief='raised', bd=2)
        generics_card.pack(fill='both', expand=True)
        tk.Label(generics_card, text="💊 Generic Alternatives", font=('Segoe UI', 16, 'bold'), fg=COLORS['text_dark'], bg=COLORS['card']).pack(pady=(15,10))
        
        canvas = tk.Canvas(generics_card, bg=COLORS['card'])
        scrollbar = ttk.Scrollbar(generics_card, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=COLORS['card'])
        scrollable_frame.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0,0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        generics_found = False
        
        for i in range(MAX_GENERICS):
            name_idx = 3 + (i * 4)
            comp_idx = 4 + (i * 4)
            price_idx = 5 + (i * 4)
            side_idx = 6 + (i * 4)
            
            if name_idx < len(data) and data[name_idx] and str(data[name_idx]).strip():
                generics_found = True
                generic_frame = tk.Frame(scrollable_frame, bg='#E8F4FD', relief='raised', bd=1)
                generic_frame.pack(fill='x', padx=15, pady=5)
                
                tk.Label(generic_frame, text=f"Generic {i+1}: {data[name_idx]}", font=('Segoe UI', 12, 'bold'), fg=COLORS['primary'], bg='#E8F4FD').grid(row=0,column=0, sticky='w', padx=10,pady=(10,5))
                
                comp_text = data[comp_idx] if comp_idx < len(data) and data[comp_idx] else "Not specified"
                tk.Label(generic_frame, text=f"Composition: {comp_text}", font=('Segoe UI',10), fg=COLORS['text_dark'], bg='#E8F4FD').grid(row=1,column=0, sticky='w', padx=10)
                
                price_text = data[price_idx] if price_idx < len(data) and data[price_idx] else "Not specified"
                tk.Label(generic_frame, text=f"💰 Price: {price_text}", font=('Segoe UI',10), fg=COLORS['success'], bg='#E8F4FD').grid(row=2,column=0, sticky='w', padx=10)
                
                side_text = data[side_idx] if side_idx < len(data) and data[side_idx] else "Not specified"
                tk.Label(generic_frame, text=f"⚠️ Side Effects: {side_text}", font=('Segoe UI',10), fg=COLORS['secondary'], bg='#E8F4FD', wraplength=800).grid(row=3,column=0, sticky='w', padx=10,pady=(0,10))
        
        if not generics_found:
            tk.Label(scrollable_frame, text="No generic alternatives found.", font=('Segoe UI', 12), fg=COLORS['text_light'], bg=COLORS['card']).pack(pady=20)
        
        canvas.pack(side='left', fill='both', expand=True, padx=(15,0), pady=(0,15))
        scrollbar.pack(side='right', fill='y', pady=(0,15), padx=(0,15))
    
    def open_medicine_ui(self):
        MedicineFormWindow(self.root, self.refresh_recent)
    
    def refresh_recent(self):
        for widget in self.recent_frame.winfo_children():
            widget.destroy()
        self.load_recent_medicines()
    
    def show_statistics(self):
        try:
            wb = load_workbook(FILE_NAME)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            
            total_medicines = len(rows)
            total_generics = 0
            for row in rows:
                for i in range(MAX_GENERICS):
                    name_idx = 3 + (i * 4)
                    if name_idx < len(row) and row[name_idx] and str(row[name_idx]).strip():
                        total_generics += 1
            wb.close()
            
            stats_window = tk.Toplevel(self.root)
            stats_window.title("📈 Database Statistics")
            stats_window.geometry("400x300")
            stats_window.configure(bg=COLORS['background'])
            stats_window.transient(self.root)
            stats_window.grab_set()
            
            tk.Label(stats_window, text="📈 Database Statistics", font=('Segoe UI',18,'bold'), fg=COLORS['text_dark'], bg=COLORS['background']).pack(pady=20)
            stats_frame = tk.Frame(stats_window, bg=COLORS['card'], relief='raised', bd=2)
            stats_frame.pack(fill='both', expand=True, padx=20, pady=(0,20))
            
            tk.Label(stats_frame, text=f"Total Medicines: {total_medicines}", font=('Segoe UI',14), fg=COLORS['text_dark'], bg=COLORS['card']).pack(pady=15)
            tk.Label(stats_frame, text=f"Total Generic Alternatives: {total_generics}", font=('Segoe UI',14), fg=COLORS['text_dark'], bg=COLORS['card']).pack(pady=15)
            avg_generics = round(total_generics/total_medicines,2) if total_medicines>0 else 0
            tk.Label(stats_frame, text=f"Average Generics per Medicine: {avg_generics}", font=('Segoe UI',14), fg=COLORS['text_dark'], bg=COLORS['card']).pack(pady=15)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error loading statistics: {str(e)}")
    
    def run(self):
        self.root.mainloop()

class MedicineFormWindow:
    def __init__(self, parent, refresh_callback):
        self.parent = parent
        self.refresh_callback = refresh_callback
        self.generic_frames = []
        
        self.window = tk.Toplevel(parent)
        self.setup_window()
        self.create_form()
        
    def setup_window(self):
        self.window.title("➕ Add New Medicine")
        self.window.geometry("1000x800")
        self.window.configure(bg=COLORS['background'])
        self.window.transient(self.parent)
        self.window.grab_set()
        
    def create_form(self):
        header_frame = tk.Frame(self.window, bg=COLORS['success'], height=60)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="➕ Add New Medicine", font=('Segoe UI',18,'bold'), fg='white', bg=COLORS['success']).pack(expand=True)
        
        main_frame = tk.Frame(self.window, bg=COLORS['background'])
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        info_card = tk.Frame(main_frame, bg=COLORS['card'], relief='raised', bd=2)
        info_card.pack(fill='x', pady=(0,15))
        tk.Label(info_card, text="Medicine Information", font=('Segoe UI',16,'bold'), fg=COLORS['text_dark'], bg=COLORS['card']).pack(pady=(15,10))
        
        tk.Label(info_card, text="Medicine Name *", font=('Segoe UI',12,'bold'), fg=COLORS['text_dark'], bg=COLORS['card']).pack(anchor='w', padx=20)
        self.entry_name = tk.Entry(info_card, font=('Segoe UI',12), width=50)
        self.entry_name.pack(padx=20, pady=(0,10))
        
        tk.Label(info_card, text="Composition *", font=('Segoe UI',12,'bold'), fg=COLORS['text_dark'], bg=COLORS['card']).pack(anchor='w', padx=20)
        self.entry_composition = tk.Entry(info_card, font=('Segoe UI',12), width=50)
        self.entry_composition.pack(padx=20, pady=(0,15))
        
        self.generics_container = tk.Frame(main_frame, bg=COLORS['background'])
        self.generics_container.pack(fill='both', expand=True)
        
        add_generic_btn = ModernButton(main_frame, "➕ Add Generic Medicine", self.add_generic, COLORS['primary'], width=25)
        add_generic_btn.pack(pady=15)
        
        self.status_label = tk.Label(main_frame, text="", font=('Segoe UI',11), bg=COLORS['background'])
        self.status_label.pack()
        
        save_btn = ModernButton(main_frame, "💾 Save Medicine", self.save_medicine, COLORS['success'], width=25)
        save_btn.pack(pady=20)
    
    def add_generic(self):
        if len(self.generic_frames) >= MAX_GENERICS:
            self.status_label.config(text=f"⚠️ Maximum {MAX_GENERICS} generics allowed!", fg='red')
            return
        
        idx = len(self.generic_frames) + 1
        generic_frame = tk.Frame(self.generics_container, bg='#F8F9FA', relief='raised', bd=1)
        generic_frame.pack(fill='x', pady=5)
        
        header_frame = tk.Frame(generic_frame, bg='#F8F9FA')
        header_frame.pack(fill='x', padx=5, pady=5)
        tk.Label(header_frame, text=f"Generic {idx}", font=('Segoe UI',12,'bold'), fg=COLORS['primary'], bg='#F8F9FA').pack(side='left')
        tk.Button(header_frame, text="❌ Remove", bg='red', fg='white', relief='flat', command=lambda f=generic_frame: self.remove_generic(f)).pack(side='right')
        
        fields_frame = tk.Frame(generic_frame, bg='#F8F9FA')
        fields_frame.pack(padx=5, pady=(0,5))
        
        tk.Label(fields_frame, text="Name:", bg='#F8F9FA', font=('Segoe UI',10,'bold')).grid(row=0,column=0, sticky='w', padx=(0,5), pady=2)
        entry_name = tk.Entry(fields_frame, font=('Segoe UI',10), width=20)
        entry_name.grid(row=0,column=1, padx=5, pady=2)
        
        tk.Label(fields_frame, text="Composition:", bg='#F8F9FA', font=('Segoe UI',10,'bold')).grid(row=0,column=2, sticky='w', padx=(10,5), pady=2)
        entry_comp = tk.Entry(fields_frame, font=('Segoe UI',10), width=20)
        entry_comp.grid(row=0,column=3, padx=5, pady=2)
        
        tk.Label(fields_frame, text="Price:", bg='#F8F9FA', font=('Segoe UI',10,'bold')).grid(row=1,column=0, sticky='w', padx=(0,5), pady=2)
        entry_price = tk.Entry(fields_frame, font=('Segoe UI',10), width=20)
        entry_price.grid(row=1,column=1, padx=5, pady=2)
        
        tk.Label(fields_frame, text="Side Effects:", bg='#F8F9FA', font=('Segoe UI',10,'bold')).grid(row=1,column=2, sticky='w', padx=(10,5), pady=2)
        entry_side = tk.Entry(fields_frame, font=('Segoe UI',10), width=20)
        entry_side.grid(row=1,column=3, padx=5, pady=2)
        
        self.generic_frames.append({
            'frame': generic_frame,
            'name': entry_name,
            'composition': entry_comp,
            'price': entry_price,
            'side_effects': entry_side
        })
    
    def remove_generic(self, frame):
        for g in self.generic_frames:
            if g['frame'] == frame:
                g['frame'].destroy()
                self.generic_frames.remove(g)
                break
        for i, g in enumerate(self.generic_frames):
            header_label = g['frame'].winfo_children()[0].winfo_children()[0]
            header_label.config(text=f"Generic {i+1}")
    
    def save_medicine(self):
        name = self.entry_name.get().strip()
        composition = self.entry_composition.get().strip()
        
        if not name or not composition:
            self.status_label.config(text="⚠️ Medicine Name and Composition are required!", fg='red')
            return
        
        data_row = [name, composition, datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        
        for g in self.generic_frames:
            data_row.extend([
                g['name'].get().strip(),
                g['composition'].get().strip(),
                g['price'].get().strip(),
                g['side_effects'].get().strip()
            ])
        
        for _ in range(MAX_GENERICS - len(self.generic_frames)):
            data_row.extend(["", "", "", ""])
        
        try:
            wb = load_workbook(FILE_NAME)
            ws = wb.active
            ws.append(data_row)
            wb.save(FILE_NAME)
            wb.close()
            
            self.status_label.config(text="✅ Medicine saved successfully!", fg='green')
            self.refresh_callback()
            self.window.destroy()
        except Exception as e:
            self.status_label.config(text=f"❌ Error saving medicine: {str(e)}", fg='red')

initialize_excel()
app = MedicineApp()
app.run()
